"""
OrangeJudge, a competitive programming platform

Copyright (C) 2024-2025 LittleOrange666 (orangeminecraft123@gmail.com)

This program is free software: you can redistribute it and/or modify
it under the terms of the GNU Affero General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

This program is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU Affero General Public License for more details.

You should have received a copy of the GNU Affero General Public License
along with this program.  If not, see <https://www.gnu.org/licenses/>.
"""
import csv
import multiprocessing
import time
import traceback
from datetime import datetime, timedelta
from io import BytesIO, TextIOWrapper
from multiprocessing import Process
from time import sleep

from cachetools import TTLCache, cached
from flask import request
from flask_login import current_user
from loguru import logger
from openpyxl.reader.excel import load_workbook
from werkzeug.datastructures import ImmutableMultiDict

from . import tools, datas, tasks, objs, server, login
from .objs import Permission, ContestStatus

actions = tools.Switcher()


def create_contest(name: str, user: datas.User) -> str:
    if len(name) > 120:
        server.custom_abort(400, "Contest name too long")
    ccnt = datas.count(datas.Contest)
    cidx = ccnt + 1
    while datas.count(datas.Contest, cid=str(cidx)) > 0:
        cidx += 1
    cid = str(cidx)
    start_time = datetime.now().replace(second=0, microsecond=0) + timedelta(days=1)
    start_timestamp = start_time.timestamp()
    info = objs.ContestData(name=name, users=[user.username], start=int(start_timestamp), elapsed=60)
    dat = datas.Contest(id=cidx, cid=cid, name=name, data=objs.as_dict(info), user=user)
    per = datas.Period(start_time=start_time,
                       end_time=start_time + timedelta(hours=1),
                       ended=False,
                       running=False,
                       contest_id=dat.id,
                       is_virtual=False)
    datas.add(dat, per)
    datas.flush()  # 重要 !!!
    dat.main_period_id = per.id
    datas.add(dat)
    dat.path.mkdir(parents=True, exist_ok=True)
    tools.write_json({}, dat.path / "standings.json")
    return cid


def calidx(idx: int) -> str:
    s = ""
    if idx >= 26:
        s = calidx(idx // 26 - 1)
        idx %= 26
    return s + chr(ord('A') + idx)


@actions.bind
def add_problem(form: ImmutableMultiDict[str, str], cdat: datas.Contest, dat: objs.ContestData) -> str:
    pid = form["pid"]
    pdat: datas.Problem = datas.first_or_404(datas.Problem, pid=pid)
    if len(pdat.datas.versions) == 0:
        server.custom_abort(409, "Problem has no available version")
    for idx, obj in dat.problems.items():
        if obj.pid == pid:
            server.custom_abort(409, "Problem already added to contest")
    idx = 0
    while calidx(idx) in dat.problems:
        idx += 1
    dat.problems[calidx(idx)] = objs.ContestProblem(pid=pid, name=pdat.name)
    return "index_page"


@actions.bind
def remove_problem(form: ImmutableMultiDict[str, str], cdat: datas.Contest, dat: objs.ContestData) -> str:
    idx = form["idx"]
    if idx not in dat.problems:
        server.custom_abort(409, "Problem not found in contest")
    del dat.problems[idx]
    return "index_page"


@actions.bind
def add_participant(form: ImmutableMultiDict[str, str], cdat: datas.Contest, dat: objs.ContestData) -> str:
    user: datas.User = datas.first_or_404(datas.User, username=form["username"].lower())
    if user.username in dat.participants:
        server.custom_abort(409, "User is already a participant")
    dat.participants.append(user.username)
    return "participants"


@actions.bind
def add_participants(form: ImmutableMultiDict[str, str], cdat: datas.Contest, dat: objs.ContestData) -> str:
    file = request.files["file"]
    if not file or file.filename == "":
        server.custom_abort(400, "No file provided")
    ext = file.filename.rsplit(".", 1)[-1].lower()
    if ext not in ("xlsx", "csv"):
        server.custom_abort(400, "Unsupported file type")
    try:
        in_memory_file = BytesIO(file.stream.read())
        if ext == ".xlsx":
            wb = load_workbook(in_memory_file, data_only=True)
            ws = wb.active
            arr = [[str(cell.value) for cell in row] for row in ws.iter_rows()]
        else:
            wrapper = TextIOWrapper(in_memory_file, encoding="utf-8")
            reader = csv.reader(wrapper)
            arr = [[cell for cell in row] for row in reader]
    except Exception as e:
        server.custom_abort(400, "Failed to read file: " + str(e))
        return "participants"
    if len(arr) < 1:
        server.custom_abort(400, "No data found in file")
    if "username" not in arr[0]:
        out = [line[0].strip() for line in arr if len(line) > 0 and line[0].strip() != ""]
    else:
        i0 = arr[0].index("username")
        out = [line[i0].strip() for line in arr if len(line) > i0 and line[i0].strip() != ""]
    if len(out) < 1:
        server.custom_abort(400, "No valid usernames found in file")
    bad_usernames = []
    for username in out:
        username = username.lower()
        if username in dat.participants:
            continue
        user: datas.User = datas.first(datas.User, username=username)
        if user is None:
            bad_usernames.append(username)
    if len(bad_usernames) > 0:
        server.custom_abort(400, "The following usernames do not exist: " + ", ".join(bad_usernames))
        return "participants"
    for username in out:
        username = username.lower()
        if username in dat.participants:
            continue
        dat.participants.append(username)
    return "participants"


@actions.bind
def remove_participant(form: ImmutableMultiDict[str, str], cdat: datas.Contest, dat: objs.ContestData) -> str:
    user: datas.User = datas.first_or_404(datas.User, username=form["username"].lower())
    if user.username not in dat.participants:
        server.custom_abort(409, "User is not a participant")
    dat.participants.remove(user.username)
    return "participants"


@actions.bind
def change_settings(form: ImmutableMultiDict[str, str], cdat: datas.Contest, dat: objs.ContestData) -> str:
    contest_title = form["contest_title"]
    if len(contest_title) < 1 or len(contest_title) > 120:
        server.custom_abort(400, "Contest title length must be between 1 and 120 characters")
    start_time = 0
    try:
        start_time = tools.to_datetime(form["start_time"], second=0, microsecond=0).timestamp()
    except ValueError:
        server.custom_abort(400, "Invalid start time format")
    if not form["elapsed_time"].isdigit():
        server.custom_abort(400, "Elapsed time must be a positive integer")
    elapsed_time = int(form["elapsed_time"])
    rule_type = form["rule_type"]
    if rule_type not in ("icpc", "ioi"):
        server.custom_abort(400, "Invalid rule type")
    pretest_type = form["pretest_type"]
    if pretest_type not in ("all", "last", "no"):
        server.custom_abort(400, "Invalid pretest type")
    practice_type = form["practice_type"]
    if practice_type not in ("no", "private", "public"):
        server.custom_abort(400, "Invalid practice type")
    register_type = form["register_type"]
    if register_type not in ("no", "yes"):
        server.custom_abort(400, "Invalid register type")
    show_standing = form["show_standing"]
    if show_standing not in ("no", "yes"):
        server.custom_abort(400, "Invalid show standing type")
    show_contest = form["show_contest"]
    if show_contest not in ("no", "yes"):
        server.custom_abort(400, "Invalid show contest type")
    if not form["freeze_time"].isdigit():
        server.custom_abort(400, "Freeze time must be a positive integer")
    freeze_time = int(form["freeze_time"])
    if not form["unfreeze_time"].isdigit():
        server.custom_abort(400, "Unfreeze time must be a positive integer")
    unfreeze_time = int(form["unfreeze_time"])
    if not form["penalty"].isdigit():
        server.custom_abort(400, "Penalty must be a positive integer")
    penalty = int(form["penalty"])
    per: datas.Period = datas.get_by_id(datas.Period, cdat.main_period_id)
    per.start_time = datetime.fromtimestamp(start_time)
    cdat.name = contest_title
    dat.name = contest_title
    dat.start = start_time
    dat.elapsed = elapsed_time
    dat.type = rule_type
    dat.pretest = pretest_type
    dat.practice = practice_type
    dat.can_register = (register_type == "yes")
    cdat.hidden = (show_contest == "no")
    dat.standing = objs.StandingsData(public=(show_standing == "yes"), start_freeze=freeze_time,
                                      end_freeze=unfreeze_time)
    dat.penalty = penalty
    return "edit"


@actions.bind
def save_order(form: ImmutableMultiDict[str, str], cdat: datas.Contest, dat: objs.ContestData) -> str:
    order = form["order"].split(",")
    if set(order) != set(dat.problems.keys()):
        server.custom_abort(400, "Invalid problem order")
    nw_dict = {}
    arr = sorted(order)
    for k1, k2 in zip(arr, order):
        nw_dict[k1] = dat.problems[k2]
    dat.problems = nw_dict
    return "index_page"


@actions.bind
def send_announcement(form: ImmutableMultiDict[str, str], cdat: datas.Contest, dat: objs.ContestData) -> str:
    if len(form["title"]) > 80:
        server.custom_abort(400, "Title too long")
    if len(form["content"]) > 1000:
        server.custom_abort(400, "Content too long")
    obj = datas.Announcement(time=datetime.now(),
                             title=form["title"],
                             content=form["content"],
                             user=current_user.data,
                             contest=cdat,
                             public=True,
                             question=False)
    datas.add(obj)
    return "index_page"


@actions.bind
def remove_announcement(form: ImmutableMultiDict[str, str], cdat: datas.Contest, dat: objs.ContestData) -> str:
    idx = tools.to_int(form["id"])
    obj: datas.Announcement = datas.get_or_404(datas.Announcement, idx)
    if obj.contest != cdat:
        server.custom_abort(404, "Announcement not found")
    datas.delete(obj)
    return "index_page"


@actions.bind
def save_question(form: ImmutableMultiDict[str, str], cdat: datas.Contest, dat: objs.ContestData) -> str:
    idx = tools.to_int(form["id"])
    obj: datas.Announcement = datas.get_or_404(datas.Announcement, idx)
    if obj.contest != cdat:
        server.custom_abort(404, "Question not found")
    reply = form["content"]
    if len(reply) > 1000:
        server.custom_abort(400, "Reply content too long")
    public = form.get("public", "no") == "yes"
    obj.reply = reply
    obj.reply_name = current_user.id
    obj.public = public
    datas.add(obj)
    return "index_page"


@actions.default
def action_not_found(*args):
    server.custom_abort(404, "Action not found")


def action(form: ImmutableMultiDict[str, str], cdat: datas.Contest):
    dat = cdat.datas
    cid = cdat.cid
    tp = actions.call(form["action"], form, cdat, dat)
    cdat.datas = dat
    datas.add(cdat)
    if form["action"] == "change_settings":
        for the_per in cdat.periods:
            the_per: datas.Period
            the_per.end_time = the_per.start_time + timedelta(minutes=dat.elapsed)
        datas.add(*cdat.periods)
    return f"/contest/{cid}#{tp}"


def check_super_access(dat: datas.Contest, user: login.User = None) -> bool:
    if user is None:
        user = current_user
    return user.is_authenticated and (
            user.has(Permission.admin) or user.id in dat.datas.users)


def check_access(dat: datas.Contest, user: login.User = None):
    per: datas.Period = datas.get_by_id(datas.Period, dat.main_period_id)
    if user is None:
        user = current_user
    info = dat.datas
    if per is None:
        server.custom_abort(409, "Main period not found")
    if check_super_access(dat, user):
        return
    if dat.hidden:
        server.custom_abort(404, "Contest not found")
    if user.is_authenticated:
        if user.id in info.participants:
            if per.is_running():
                return
            if per.is_over() and info.practice != objs.PracticeType.no:
                return
    if info.practice == objs.PracticeType.public and per.is_over():
        return
    server.custom_abort(403, "You do not have access to this contest")


def check_status(dat: datas.Contest, user: login.User = None) -> tuple[ContestStatus, int, bool]:
    """
    Determine the current status of a contest for a given user.

    This function evaluates the contest's status based on the user's participation,
    the contest's period, and the practice type. It handles both regular and virtual
    participants and returns the appropriate contest status, a timestamp, and a boolean
    indicating whether the user has access.

    Args:
        dat (datas.Contest): The contest object to check the status for.
        user (login.User, optional): The user whose status is being checked. Defaults to the current user.

    Returns:
        tuple[ContestStatus, int, bool]: A tuple containing:
            - ContestStatus: The current status of the contest.
            - int: A timestamp related to the status (e.g., start or end time).
            - bool: Whether the user has access to the contest.

    Raises:
        409: If a virtual period for the user is not found.
    """
    per = datas.get_or_404(datas.Period, dat.main_period_id)
    info = dat.datas
    if user is None:
        user = current_user
    if user.is_authenticated:
        if user.id in info.virtual_participants:
            vir_per: datas.Period = datas.get_by_id(datas.Period, info.virtual_participants[user.id])
            if vir_per is None:
                server.custom_abort(409, "Virtual period not found for user")
            if not vir_per.is_started():
                return ContestStatus.waiting_virtual, vir_per.start_time.timestamp(), False
            if vir_per.is_running():
                return ContestStatus.running_virtual, vir_per.end_time.timestamp(), True
        if user.id in info.participants:
            if not per.is_started():
                return ContestStatus.waiting, per.start_time.timestamp(), False
            if per.is_running():
                return ContestStatus.running, per.end_time.timestamp(), True
            if per.is_over() and info.practice != objs.PracticeType.no:
                return ContestStatus.practice, 0, True
        if check_super_access(dat):
            return ContestStatus.testing, 0, True
        if per.is_over() and info.practice == objs.PracticeType.public:
            return ContestStatus.practice, 0, True
    if info.practice == objs.PracticeType.public and per.is_over():
        return ContestStatus.guest, 0, True
    return ContestStatus.guest, 0, False


def check_period(dat: datas.Contest, user: login.User = None) -> int:
    if user is None:
        user = current_user
    main_per = datas.get_or_404(datas.Period, dat.main_period_id)
    info = dat.datas
    if user.id in info.participants and main_per.is_running():
        return dat.main_period_id
    if user.id in info.virtual_participants:
        per_id = info.virtual_participants[user.id]
        cur_per = datas.get_or_404(datas.Period, per_id)
        if cur_per.is_running():
            return per_id
    return 0


standing_lock = multiprocessing.Lock()


@cached(cache=TTLCache(maxsize=20, ttl=10), lock=standing_lock)
def get_standing(cid: str):
    cdat = datas.first_or_404(datas.Contest, cid=cid)
    ret = []
    mp = {}
    rmp = {}
    info = cdat.datas
    for k, v in info.problems.items():
        rmp[v.pid] = k
    for dat in cdat.submissions.filter_by(completed=True).all():
        dat: datas.Submission
        res = dat.results
        if dat.user_id not in mp:
            mp[dat.user_id] = dat.user.display_name
        scores = {k: v.gained_score for k, v in res.group_results.items()}
        ret.append({"user": mp[dat.user_id],
                    "pid": rmp[dat.pid],
                    "scores": scores,
                    "total_score": res.total_score,
                    "time": dat.time.timestamp(),
                    "pretest": dat.just_pretest,
                    "per": dat.period_id})
    pers = []
    for per in cdat.periods:
        per: datas.Period
        pers.append({"start_time": per.start_time.timestamp(),
                     "judging": per.judging,
                     "idx": per.id})
    return {"submissions": ret,
            "rule": info.type.name,
            "pids": list(info.problems.keys()),
            "penalty": info.penalty,
            "pers": pers,
            "main_per": cdat.main_period_id,
            "participants": info.participants,
            "virtual_participants": info.virtual_participants}


def reject(dat: datas.Submission):
    dat.simple_result = "ignored"
    dat.simple_result_flag = objs.TaskResult.SKIP.name
    res = dat.results
    res.total_score = 0
    for k in res.group_results:
        res.group_results[k].gained_score = 0
    dat.results = res


def contest_worker():
    while True:
        try:
            with datas.SessionContext():
                for dat in datas.get_all(datas.Period, running=True):
                    if dat.is_over():
                        dat.running = False
                        dat.ended = True
                        cdat: datas.Contest = dat.contest
                        pretest = cdat.datas.pretest
                        if pretest != objs.PretestType.no:
                            submissions = dat.submissions.filter_by(just_pretest=True).all()
                            dic: dict[tuple[int, str], datas.Submission] = {}
                            datas.add(*submissions)
                            for submission in submissions:
                                submission: datas.Submission
                                submission.just_pretest = False
                                key = (submission.user_id, submission.pid)
                                if submission.simple_result.lower() not in ("je", "ce"):
                                    reject(submission)
                                    if pretest == objs.PretestType.all:
                                        tasks.rejudge(submission)
                                    else:
                                        dic[key] = submission
                            if pretest == objs.PretestType.last:
                                for v in dic.values():
                                    tasks.rejudge(v)
                            datas.add(*submissions)
                    datas.add(dat)
            sleep(5)
            with datas.SessionContext():
                for dat in datas.get_all(datas.Period, running=False, ended=False):
                    dat: datas.Period
                    if dat.is_running():
                        dat.running = True
                        dat.judging = True
                        datas.add(dat)
            sleep(5)
            with datas.SessionContext():
                for dat in datas.get_all(datas.Period, running=False, ended=True, judging=True):
                    dat: datas.Period
                    if dat.submissions.filter_by(completed=False).count() == 0:
                        dat.judging = False
                        datas.add(dat)
            sleep(5)
            with datas.SessionContext():
                for dat in datas.get_all(datas.Period, running=False, ended=True, judging=False):
                    dat: datas.Period
                    if not dat.is_started():
                        dat.ended = False
                        datas.add(dat)
            sleep(5)
        except Exception as e:
            logger.error(f"Error in contest worker: {e}")
            logger.debug(traceback.format_exc())
            time.sleep(60)


def init():
    Process(target=contest_worker).start()
