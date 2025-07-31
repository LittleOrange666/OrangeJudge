"""
OrangeJudge, a competitive programming platform

Copyright (C) 2024-2025 LittleOrange666 (orangeminecraft123@gmail.com)

This program is free software: you can redistribute it and/or modify
it under the terms of the GNU Affero General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

This program is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU Affero General Public License for more details.

You should have received a copy of the GNU Affero General Public License
along with this program.  If not, see <https://www.gnu.org/licenses/>.
"""
import csv
import json
import os
import traceback
import uuid
from io import BytesIO, TextIOWrapper

from flask import abort, render_template, redirect, request, jsonify
from flask_login import login_required, current_user
from limits import parse
from openpyxl import load_workbook
from pygments import highlight, lexers
from pygments.formatters import HtmlFormatter
from werkzeug.utils import secure_filename

from .. import tools, server, constants, executing, tasks, datas, contests, login, config, objs
from ..constants import problem_path, preparing_problem_path
from ..objs import Permission
from ..server import sending_file

app = server.app


def update_user():
    user: datas.User = datas.first_or_404(datas.User, username=request.form["username"])
    user.display_name = request.form["display_name"]
    if len(request.form["password"]) > 1:
        user.password_sha256_hex = login.try_hash(request.form["password"])
    perms = user.permission_list()
    new_perms = request.form["permissions"].split(";")
    for perm_name in ("admin", "make_problems"):
        if perm_name in new_perms:
            if perm_name not in perms:
                perms.append(perm_name)
        else:
            if perm_name in perms:
                perms.remove(perm_name)
    user.permissions = ";".join(perms)
    datas.add(user)
    return "OK", 200


def parse_user():
    file = request.files.get("file")
    if not file or file.filename == "":
        return "No file provided", 400
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in (".xlsx", ".csv"):
        return "Unsupported file type", 400
    try:
        in_memory_file = BytesIO(file.stream.read())
        if ext == ".xlsx":
            wb = load_workbook(in_memory_file, data_only=True)
            ws = wb.active
            arr = [[str(cell.value) for cell in row] for row in ws.iter_rows()]
        else:
            wrapper = TextIOWrapper(in_memory_file, encoding="utf-8")
            reader = csv.reader(wrapper)
            arr = [[cell for cell in row] for row in reader]
    except Exception as e:
        traceback.print_exception(e)
        return "Failed to read file", 400
    if len(arr) < 1:
        return "No data found", 400
    title = [cell.lower() for cell in arr[0]]
    i0 = 0
    i1 = 1
    i2 = 2
    i3 = 3
    if "username" in title and "password" in title:
        i0 = title.index("username")
        i1 = title.index("password")
        if "email" in title:
            i2 = title.index("email")
        else:
            i2 = -1
        if "display_name" in title:
            i3 = title.index("display_name")
        else:
            i3 = -1
        arr = arr[1:]
    out = []
    existed = []
    for row in arr:
        if len(row) < 2:
            continue
        username = row[i0].strip().lower() if len(row) > i0 else ""
        password = row[i1].strip() if len(row) > i1 else ""
        email = row[i2].strip() if 0 <= i2 < len(row) else ""
        display_name = row[i3].strip() if 0 <= i3 < len(row) else username
        if not username or not password:
            continue
        if datas.count(datas.User, username=username) > 0 or (email and datas.count(datas.User, email=email) > 0):
            existed.append([username, email])
            continue
        out.append([username, password, email, display_name])
    if len(out) < 1:
        return "No valid user data found", 400
    return jsonify({"users": out, "existed": existed}), 200


def create_users():
    dat = request.form["users"]
    try:
        users = json.loads(dat)
    except json.JSONDecodeError:
        return "Invalid JSON data", 400
    if not isinstance(users, list) or len(users) < 1:
        return "No valid user data found", 400
    for user in users:
        if len(user) < 2:
            continue
        username = user[0].strip().lower()
        password = user[1].strip()
        email = user[2].strip() if len(user) > 2 else uuid.uuid5(uuid.NAMESPACE_DNS, username).hex + "@placeholder.com"
        display_name = user[3].strip() if len(user) > 3 else username
        if not username or not password:
            continue
        if datas.count(datas.User, username=username) > 0 or (email and datas.count(datas.User, email=email) > 0):
            continue
        login.create_account(email, username, password, display_name)
    return "OK", 200


def is_valid_limit_string(limit_str: str) -> bool:
    try:
        parse(limit_str)
        return True
    except Exception:
        return False


def update_config():
    config_fields = config.get_fields()
    dat_str = request.form["config"]
    try:
        dat = json.loads(dat_str)
        if not isinstance(dat, list):
            return "JSON格式錯誤", 400
        ks = [o[0] for o in dat]
        mp = {o[0]: o[1] for o in dat}
    except json.JSONDecodeError | TypeError | ValueError:
        return "JSON格式錯誤", 400

    for category in config_fields:
        for slot in category["slots"]:
            key = f"config_{category['name']}_{slot['name']}"
            if slot['type'] == "limits":
                true_keys = [k for k in ks if k.startswith(key + "_")]
                if len(true_keys) == 0:
                    return f"缺少設定項目: {key}", 400
                for true_key in true_keys:
                    val = mp[true_key]
                    str_val = f"{val[0]} per {val[1]} {val[2]}"
                    if not is_valid_limit_string(str_val):
                        return f"{str_val!r} 不是合法的速率限制", 400
            else:
                if key not in ks:
                    return f"缺少設定項目: {key}", 400
                if slot['type'] == "bool":
                    if mp[key] not in ("true", "false"):
                        return f"{mp[key]!r} 不是合法的布林值", 400
                    mp[key] = True if mp[key] == "true" else False
                elif slot['type'] == "int":
                    if not mp[key].isdigit():
                        return f"{mp[key]!r} 不是合法的整數", 400
                    mp[key] = int(mp[key])
                elif slot['type'] == "limit":
                    val = mp[key]
                    str_val = f"{val[0]} per {val[1]} {val[2]}"
                    if not is_valid_limit_string(str_val):
                        return f"{str_val!r} 不是合法的速率限制", 400

    for category in config_fields:
        cat_obj = getattr(config.config, category["name"])
        for slot in category["slots"]:
            key = f"config_{category['name']}_{slot['name']}"
            if slot['type'] == "limits":
                true_keys = [k for k in ks if k.startswith(key + "_")]
                vals = []
                for true_key in true_keys:
                    val = mp[true_key]
                    str_val = f"{val[0]} per {val[1]} {val[2]}"
                    vals.append(str_val)
                setattr(cat_obj, slot['name'], vals)
            elif slot['type'] == "limit":
                val = mp[key]
                str_val = f"{val[0]} per {val[1]} {val[2]}"
                setattr(cat_obj, slot['name'], str_val)
            else:
                setattr(cat_obj, slot['name'], mp[key])
    config.save_config()
    return "OK", 200


@app.route('/admin', methods=['GET', 'POST'])
@login_required
def admin():
    if not current_user.has(Permission.root):
        abort(403)
    if request.method == 'GET':
        users = datas.query(datas.User).all()
        config_fields = config.get_fields()
        return render_template("admin.html", users=users, config_fields=config_fields,
                               enumerate=enumerate)
    else:
        if request.form["action"] == "update_user":
            return update_user()
        if request.form["action"] == "parse_user":
            return parse_user()
        if request.form["action"] == "create_users":
            return create_users()
        if request.form["action"] == "update_config":
            return update_config()
        abort(404)
